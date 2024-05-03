import os
from copy import deepcopy
from datetime import datetime

import openpyxl

from csg.config import *
from csg.dataLoader import get_courses_data, get_duties_data, get_template


class Scheduler:
    def __init__(self) -> None:
        self.dp = ""
        self.op = ""
        self.class_schedule: dict[int, list[list[str | None]]] = {}
        self.teacher_schedule: dict[str, list[list[tuple[int, str] | str]]] = {}
        self.missing_items: list[tuple[str, str, int, int]] = []

    def set_input_path(self, data_path: str):
        if os.path.exists(data_path):
            self.dp = data_path
        else:
            raise Exception("输入路径不存在")

    def set_output_path(self, output: str):
        if os.path.exists(output):
            self.op = output
        else:
            raise Exception("输出路径不存在")

    def get_sheet_by_index(self, index: int):
        wb_obj = openpyxl.load_workbook(self.dp)
        sheetnames = wb_obj.sheetnames
        return wb_obj[sheetnames[index]]

    def get_template_data(self):
        return get_template(
            self.get_sheet_by_index(SHEETS.get("TEMPLATE_INDEX"))  # type: ignore
        )

    def get_teacher_schedule_tempalte(self):
        original_template = self.get_template_data()
        result: list[list[str | tuple[int, str]]] = deepcopy(original_template)  # type: ignore
        for r_index, row in enumerate(original_template):
            for c_index, _ in enumerate(row):
                result[r_index][c_index] = ""
        return result

    def get_duties_data(self):
        return get_duties_data(self.get_sheet_by_index(SHEETS.get("DUTIES_SHEET")))  # type: ignore

    def get_courses_data(self):
        return get_courses_data(self.get_sheet_by_index(SHEETS.get("COURSE_SHEET_INDEX")))  # type: ignore

    def get_ordered_courses_items(self):
        results: list[tuple[str, str, set[int], int, int]] = []
        courses_data = self.get_courses_data()
        duties_data = self.get_duties_data()
        for k, v in duties_data.items():
            course, teacher = k
            class_ids = v
            zc, priority = courses_data[course]
            results.append((course, teacher, class_ids, zc, priority))
        return sorted(results, key=lambda x: x[4])

    def check_template(self):
        template_schedule = self.get_template_data()
        courses_data = self.get_courses_data()
        number_of_available_node = 0
        for i in template_schedule:
            for j in i:
                if j is None:
                    number_of_available_node += 1
        course_cout = 0
        for _, v in courses_data.items():
            cout, _ = v
            course_cout += cout
        if number_of_available_node == course_cout:
            return True
        else:
            return False

    def save_teacher_schedule(self):
        wb = openpyxl.Workbook()
        for teacher, schedule in self.teacher_schedule.items():
            sheet = wb.create_sheet(f"{teacher} 课表")
            header = ["节次", "星期一", "星期二", "星期三", "星期四", "星期五"]
            sheet.append(header)
            for i_row, row in enumerate(schedule):
                row = [f"{i[0]}/{i[1]}" if isinstance(i, tuple) else i for i in row]
                row.insert(0, f"第 {i_row + 1} 节")
                sheet.append(row)
        wb.remove(wb[wb.sheetnames[0]])
        filename = self.op + "/" + datetime.now().isoformat()[0:10] + "教师课表"
        for i in range(1, 100):
            if not os.path.exists(f"{filename}.xlsx"):
                wb.save(f"{filename}.xlsx")
                return
            if not os.path.exists(f"{filename}({i+1}).xlsx"):
                wb.save(f"{filename}({i+1}).xlsx")
                return

    def save_class_scheudle(self):
        wb = openpyxl.Workbook()
        for class_id, schedule in self.class_schedule.items():
            sheet = wb.create_sheet(f"{class_id} 班课表")
            header = ["节次", "星期一", "星期二", "星期三", "星期四", "星期五"]
            sheet.append(header)  # type: ignore
            for i_row, row in enumerate(schedule):
                row.insert(0, f"第 {i_row + 1} 节")
                sheet.append(row)  # type: ignore
        wb.remove(wb[wb.sheetnames[0]])
        filename = self.op + "/" + datetime.now().isoformat()[0:10] + "年级课表"
        for i in range(1, 100):
            if not os.path.exists(f"{filename}.xlsx"):
                wb.save(f"{filename}.xlsx")
                return
            if not os.path.exists(f"{filename}({i+1}).xlsx"):
                wb.save(f"{filename}({i+1}).xlsx")
                return

    def process(self):
        template_schedule = self.get_template_data()
        weekly_class_days = len(template_schedule[0])
        number_of_classes_per_day = len(template_schedule)
        ordered_items = self.get_ordered_courses_items()
        for ordered_item in ordered_items:
            course, teacher, classes, number_of_classes_per_week, priority = (
                ordered_item
            )
            number_of_remaining_courses = {
                class_id: number_of_classes_per_week for class_id in classes
            }
            if not self.teacher_schedule.get(teacher):
                self.teacher_schedule[teacher] = deepcopy(
                    self.get_teacher_schedule_tempalte()
                )
            for class_id in classes:
                if not number_of_remaining_courses[class_id] > 0:
                    continue
                if not self.class_schedule.get(class_id):
                    self.class_schedule[class_id] = deepcopy(self.get_template_data())
                for section in range(number_of_classes_per_day):
                    if not number_of_remaining_courses[class_id] > 0:
                        break
                    for day in range(weekly_class_days):
                        if not number_of_remaining_courses[class_id] > 0:
                            break
                        condition_1 = (
                            self.class_schedule[class_id][section][day] is None
                        )
                        condition_2 = (
                            (self.class_schedule[class_id][section - 1][day] != course)
                            if section > 0
                            else True
                        )
                        condition_3 = self.teacher_schedule[teacher][section][day] == ""
                        condition_4 = (
                            (
                                self.teacher_schedule[teacher][section - 1][day][0]
                                != class_id
                            )
                            if (section > 0)
                            and (priority < CONTINUOUS_BOUNDARY)
                            and isinstance(
                                self.teacher_schedule[teacher][section - 1][day], tuple
                            )
                            else True
                        )
                        if condition_1 and condition_2 and condition_3 and condition_4:
                            self.class_schedule[class_id][section][day] = course
                            self.teacher_schedule[teacher][section][day] = (
                                class_id,
                                course,
                            )
                            number_of_remaining_courses[class_id] -= 1
                        if (section == number_of_classes_per_day - 1) and (
                            day == weekly_class_days - 1
                        ):
                            self.missing_items.append(
                                (
                                    course,
                                    teacher,
                                    class_id,
                                    number_of_remaining_courses[class_id],
                                )
                            )
